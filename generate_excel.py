"""
Restaurant Data Excel Generator
Generates a beautifully formatted Excel file with Google Maps links and phone number slots.
"""

import csv
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


def create_google_maps_link(address: str, name: str) -> str:
    """Generate a Google Maps search URL for the given address and restaurant name."""
    search_query = f"{name}, {address}, București, Romania"
    encoded_query = urllib.parse.quote(search_query)
    return f"https://www.google.com/maps/search/?api=1&query={encoded_query}"


def read_csv_data(filepath: str) -> list:
    """Read restaurant data from CSV file."""
    data = []
    with open(filepath, 'r', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            if len(row) >= 4:
                data.append({
                    'zone': row[0].strip(),
                    'name': row[1].strip(),
                    'address': row[2].strip(),
                    'type': row[3].strip()
                })
    return data


def create_excel(data: list, output_path: str):
    """Create a beautifully formatted Excel file from the restaurant data."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Restaurants"
    
    # Define colors
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    zone_fills = {
        '1': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # Light green
        '2': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),  # Light blue
        '3': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # Light orange
        '4': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # Light yellow
        '5': PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),  # Light purple
        '6': PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),  # Peach
    }
    default_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    thick_border = Border(
        left=Side(style='medium', color='2F5496'),
        right=Side(style='medium', color='2F5496'),
        top=Side(style='medium', color='2F5496'),
        bottom=Side(style='medium', color='2F5496')
    )
    
    # Define fonts
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    title_font = Font(name='Calibri', size=18, bold=True, color='2F5496')
    normal_font = Font(name='Calibri', size=11)
    link_font = Font(name='Calibri', size=11, color='0563C1', underline='single')
    zone_font = Font(name='Calibri', size=11, bold=True)
    
    # Add title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "🍽️ Restaurant Directory - București"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    # Add subtitle
    ws.merge_cells('A2:F2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = "Click on addresses to open in Google Maps • Add phone numbers in the Phone column"
    subtitle_cell.font = Font(name='Calibri', size=10, italic=True, color='666666')
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    # Empty row for spacing
    ws.row_dimensions[3].height = 10
    
    # Headers (row 4)
    headers = ['Zone', 'Restaurant Name', 'Address', 'Type', 'Google Maps', 'Phone Number']
    header_row = 4
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thick_border
    
    ws.row_dimensions[header_row].height = 30
    
    # Set column widths
    column_widths = {
        'A': 8,    # Zone
        'B': 32,   # Restaurant Name
        'C': 40,   # Address
        'D': 22,   # Type
        'E': 18,   # Google Maps
        'F': 18,   # Phone Number
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Add data rows
    current_zone = None
    row_num = header_row + 1
    
    for item in data:
        zone = item['zone']
        name = item['name']
        address = item['address']
        rest_type = item['type']
        maps_link = create_google_maps_link(address, name)
        
        # Get fill color based on zone
        fill = zone_fills.get(zone, default_fill)
        
        # Zone column
        zone_cell = ws.cell(row=row_num, column=1, value=zone)
        zone_cell.font = zone_font
        zone_cell.fill = fill
        zone_cell.alignment = Alignment(horizontal='center', vertical='center')
        zone_cell.border = thin_border
        
        # Restaurant Name column
        name_cell = ws.cell(row=row_num, column=2, value=name)
        name_cell.font = Font(name='Calibri', size=11, bold=True)
        name_cell.fill = fill
        name_cell.alignment = Alignment(horizontal='left', vertical='center')
        name_cell.border = thin_border
        
        # Address column (clickable link to Google Maps)
        address_cell = ws.cell(row=row_num, column=3, value=address)
        address_cell.hyperlink = maps_link
        address_cell.font = link_font
        address_cell.fill = fill
        address_cell.alignment = Alignment(horizontal='left', vertical='center')
        address_cell.border = thin_border
        
        # Type column
        type_cell = ws.cell(row=row_num, column=4, value=rest_type)
        type_cell.font = normal_font
        type_cell.fill = fill
        type_cell.alignment = Alignment(horizontal='center', vertical='center')
        type_cell.border = thin_border
        
        # Google Maps button column
        maps_cell = ws.cell(row=row_num, column=5, value="📍 Open Map")
        maps_cell.hyperlink = maps_link
        maps_cell.font = Font(name='Calibri', size=10, color='0563C1', underline='single')
        maps_cell.fill = fill
        maps_cell.alignment = Alignment(horizontal='center', vertical='center')
        maps_cell.border = thin_border
        
        # Phone Number column (empty for user to fill)
        phone_cell = ws.cell(row=row_num, column=6, value="")
        phone_cell.font = normal_font
        phone_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        phone_cell.alignment = Alignment(horizontal='center', vertical='center')
        phone_cell.border = thin_border
        
        ws.row_dimensions[row_num].height = 22
        row_num += 1
    
    # Add zone summary at the bottom
    row_num += 2
    ws.merge_cells(f'A{row_num}:F{row_num}')
    summary_cell = ws.cell(row=row_num, column=1, value="📊 Zone Color Legend")
    summary_cell.font = Font(name='Calibri', size=12, bold=True, color='2F5496')
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_num += 1
    zone_names = {
        '1': 'Sector 1',
        '2': 'Sector 2', 
        '3': 'Sector 3',
        '4': 'Sector 4',
        '5': 'Sector 5',
        '6': 'Sector 6'
    }
    
    # Get unique zones from data
    unique_zones = sorted(set(item['zone'] for item in data))
    
    for zone in unique_zones:
        zone_cell = ws.cell(row=row_num, column=1, value=zone)
        zone_cell.fill = zone_fills.get(zone, default_fill)
        zone_cell.font = zone_font
        zone_cell.alignment = Alignment(horizontal='center', vertical='center')
        zone_cell.border = thin_border
        
        name_cell = ws.cell(row=row_num, column=2, value=zone_names.get(zone, f'Zone {zone}'))
        name_cell.fill = zone_fills.get(zone, default_fill)
        name_cell.font = normal_font
        name_cell.alignment = Alignment(horizontal='left', vertical='center')
        name_cell.border = thin_border
        
        # Count restaurants in this zone
        count = sum(1 for item in data if item['zone'] == zone)
        count_cell = ws.cell(row=row_num, column=3, value=f"{count} restaurants")
        count_cell.fill = zone_fills.get(zone, default_fill)
        count_cell.font = Font(name='Calibri', size=11, italic=True)
        count_cell.alignment = Alignment(horizontal='left', vertical='center')
        count_cell.border = thin_border
        
        row_num += 1
    
    # Freeze panes (freeze header row)
    ws.freeze_panes = 'A5'
    
    # Add auto-filter
    ws.auto_filter.ref = f"A4:F{header_row + len(data)}"
    
    # Save the workbook
    wb.save(output_path)
    print(f"✅ Excel file created successfully: {output_path}")
    print(f"📊 Total restaurants: {len(data)}")
    print(f"🗺️ Unique zones: {len(unique_zones)}")


def main():
    csv_path = "data.csv"
    excel_path = "Restaurants_Directory.xlsx"
    
    print("🍽️ Restaurant Excel Generator")
    print("=" * 40)
    
    # Read data
    print("📖 Reading CSV data...")
    data = read_csv_data(csv_path)
    
    if not data:
        print("❌ No data found in CSV file!")
        return
    
    print(f"   Found {len(data)} restaurants")
    
    # Create Excel
    print("📝 Creating Excel file...")
    create_excel(data, excel_path)
    
    print("=" * 40)
    print("🎉 Done! Open the Excel file to view your restaurant directory.")
    print("   - Click on addresses or '📍 Open Map' to open Google Maps")
    print("   - Add phone numbers in the 'Phone Number' column")
    print("   - Use filters to sort by zone or type")


if __name__ == "__main__":
    main()
